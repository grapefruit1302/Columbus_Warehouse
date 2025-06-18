import logging
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import xlrd
from openpyxl import load_workbook
import io

_logger = logging.getLogger(__name__)

class StockReceiptSerialWizard(models.TransientModel):
    _name = 'stock.receipt.serial.wizard'
    _description = 'Введення серійних номерів'

    receipt_id = fields.Many2one('stock.receipt.incoming', 'Документ', required=True)
    selected_line_id = fields.Many2one('stock.receipt.incoming.line', 'Позиція',
                                     required=True, domain="[('receipt_id', '=', receipt_id), ('tracking_serial', '=', True)]")
    serial_line_ids = fields.One2many('stock.receipt.serial.wizard.serial', 'wizard_id', 'Серійні номери')
    selected_product_name = fields.Char('Обраний товар', related='selected_line_id.product_name', readonly=True)
    selected_qty = fields.Float('Необхідна кількість', related='selected_line_id.qty', readonly=True)
    current_serial_count = fields.Integer('Введено S/N', compute='_compute_current_serial_count')
    warning_message = fields.Char('Попередження', readonly=True)

    @api.depends('serial_line_ids')
    def _compute_current_serial_count(self):
        for wizard in self:
            wizard.current_serial_count = len([line for line in wizard.serial_line_ids if line.serial_number])

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        receipt_id = self.env.context.get('default_receipt_id')
        line_id = self.env.context.get('default_selected_line_id')
        warning_message = self.env.context.get('warning_message')

        if receipt_id and line_id:
            res['receipt_id'] = receipt_id
            res['selected_line_id'] = line_id
            if warning_message:
                res['warning_message'] = warning_message
            line = self.env['stock.receipt.incoming.line'].browse(line_id)
            if line.serial_numbers:
                serials = [s.strip() for s in line.serial_numbers.splitlines() if s.strip()]
                res['serial_line_ids'] = [(0, 0, {'serial_number': serial}) for serial in serials]
        
        return res
    
    def action_load_from_file(self):
        if not self.selected_line_id:
            raise UserError('Спочатку оберіть товар для введення серійних номерів!')
            
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stock.receipt.serial.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_parent_wizard_id': self.id,
                'default_selected_line_id': self.selected_line_id.id,
            }
        }
    
    def _get_duplicate_serials(self):
        """Повертає список дублікатів серійних номерів"""
        serials = [line.serial_number for line in self.serial_line_ids if line.serial_number]
        duplicates = [s for s in serials if serials.count(s) > 1]
        return list(set(duplicates))
    
    def _validate_serial_numbers(self):
        serials = [line.serial_number for line in self.serial_line_ids if line.serial_number]
        
        duplicates = self._get_duplicate_serials()
        if duplicates:
            # Відкриваємо візард для обробки дублікатів
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'stock.receipt.serial.duplicate.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_parent_wizard_id': self.id,
                    'default_duplicates': ', '.join(duplicates),
                }
            }
        
        if self.selected_line_id and len(serials) > int(self.selected_line_id.qty):
            raise ValidationError(
                f'Кількість серійних номерів ({len(serials)}) не може перевищувати кількість товару ({int(self.selected_line_id.qty)})'
            )
        
        return True
    
    def action_save_and_close(self):
        self.ensure_one()
        validation_result = self._validate_serial_numbers()
        
        # Якщо валідація повернула дію (візард дублікатів), виконуємо її
        if isinstance(validation_result, dict) and validation_result.get('type') == 'ir.actions.act_window':
            return validation_result
        
        # Якщо валідація пройшла успішно, зберігаємо
        serials = [line.serial_number for line in self.serial_line_ids if line.serial_number]
        self.selected_line_id.serial_numbers = '\n'.join(serials)
        return {'type': 'ir.actions.act_window_close'}

    def remove_duplicates(self):
        """Видаляє дублікати серійних номерів"""
        serials = []
        unique_serials = []
        
        for line in self.serial_line_ids:
            if line.serial_number:
                if line.serial_number not in serials:
                    serials.append(line.serial_number)
                    unique_serials.append(line.serial_number)
        
        # Очищаємо і створюємо нові записи без дублікатів
        self.serial_line_ids = [(5, 0, 0)]  # Видаляємо всі існуючі записи
        serial_lines = [(0, 0, {'serial_number': serial}) for serial in unique_serials]
        self.serial_line_ids = serial_lines
        
        return len(serials) - len(unique_serials)  # Повертаємо кількість видалених дублікатів

class StockReceiptSerialWizardLine(models.TransientModel):
    _name = 'stock.receipt.serial.wizard.line'
    _description = 'Лінія товарів з серійними номерами'
 
    wizard_id = fields.Many2one('stock.receipt.serial.wizard', 'Wizard', required=True, ondelete='cascade')
    line_id = fields.Many2one('stock.receipt.incoming.line', 'Позиція', required=True)
    product_name = fields.Char('Модель', related='line_id.product_name', readonly=True)
    qty = fields.Float('К-ть', related='line_id.qty', readonly=True)
    current_serial_count = fields.Integer('Введено S/N', help='Поточна кількість введених S/N')
    
    def action_select_line(self):
        return self.wizard_id.action_select_line(self.line_id)

class StockReceiptSerialWizardSerial(models.TransientModel):
    _name = 'stock.receipt.serial.wizard.serial'
    _description = 'Серійний номер'
    
    wizard_id = fields.Many2one('stock.receipt.serial.wizard', 'Wizard', required=True, ondelete='cascade')
    serial_number = fields.Char('Серійний номер', required=True)

class StockReceiptSerialDuplicateWizard(models.TransientModel):
    _name = 'stock.receipt.serial.duplicate.wizard'
    _description = 'Обробка дублікатів серійних номерів'
    
    parent_wizard_id = fields.Many2one('stock.receipt.serial.wizard', 'Основний wizard', required=True)
    duplicates = fields.Char('Дублікати', readonly=True)
    
    def action_remove_duplicates(self):
        """Видаляє дублікати та повертає до основного візарда"""
        self.ensure_one()
        
        removed_count = self.parent_wizard_id.remove_duplicates()
        
        # Повертаємо до основного візарда з повідомленням
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stock.receipt.serial.wizard',
            'view_mode': 'form',
            'res_id': self.parent_wizard_id.id,
            'target': 'new',
            'context': {
                'default_warning_message': f'Видалено {removed_count} дублікатів серійних номерів. Тепер можна зберегти дані.',
            }
        }

class StockReceiptSerialImportWizard(models.TransientModel):
    _name = 'stock.receipt.serial.import.wizard'
    _description = 'Імпорт серійних номерів'
    
    parent_wizard_id = fields.Many2one('stock.receipt.serial.wizard', 'Основний wizard')
    selected_line_id = fields.Many2one('stock.receipt.incoming.line', 'Обрана позиція')
    file_data = fields.Binary('Excel файл', required=True)
    file_name = fields.Char('Назва файлу')
    
    def action_import(self):
        """Імпортує серійні номери з файлу (.xls або .xlsx)"""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError('Будь ласка, оберіть файл!')
        
        try:
            file_content = base64.b64decode(self.file_data)
            file_name = self.file_name.lower() if self.file_name else ''

            serials = []
            if file_name.endswith('.xls'):
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                for row in range(sheet.nrows):
                    cell_value = sheet.cell_value(row, 0)
                    if cell_value:
                        if isinstance(cell_value, (int, float)) and cell_value.is_integer():
                            serials.append(str(int(cell_value)))
                        else:
                            serials.append(str(cell_value).strip())
            elif file_name.endswith('.xlsx'):
                workbook = load_workbook(filename=io.BytesIO(file_content))
                sheet = workbook.active
                for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                    cell_value = row[0]
                    if cell_value:
                        if isinstance(cell_value, (int, float)) and cell_value.is_integer():
                            serials.append(str(int(cell_value)))
                        else:
                            serials.append(str(cell_value).strip())
            else:
                raise UserError('Підтримуються лише файли форматів .xls або .xlsx!')

            serials = [s for s in serials if s]
            existing_serials = [line.serial_number for line in self.parent_wizard_id.serial_line_ids]
            all_serials = existing_serials + serials
            unique_serials = list(dict.fromkeys(all_serials))
            
            self.parent_wizard_id.serial_line_ids = [(5, 0, 0)]
            serial_lines = [(0, 0, {'serial_number': serial}) for serial in unique_serials]
            self.parent_wizard_id.serial_line_ids = serial_lines
            
            action = {
                'type': 'ir.actions.act_window',
                'res_model': 'stock.receipt.serial.wizard',
                'view_mode': 'form',
                'res_id': self.parent_wizard_id.id,
                'target': 'new',
                'context': self.env.context,
            }
            
            max_qty = int(self.selected_line_id.qty) if self.selected_line_id else 0
            if len(unique_serials) > max_qty:
                self.parent_wizard_id.warning_message = (
                    f'Імпортовано {len(unique_serials)} серійних номерів, '
                    f'але кількість товару становить {max_qty}. '
                    'Будь ласка, видаліть зайві серійні номери перед збереженням.'
                )
            
            return action
        
        except Exception as e:
            raise UserError(f'Помилка читання файлу: {str(e)}')