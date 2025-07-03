# stock_balance_management/models/stock_balance_serial_detailed_wizard.py

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

class StockBalanceSerialDetailedWizard(models.TransientModel):
    _name = 'stock.balance.serial.detailed.wizard'
    _description = 'Детальний перегляд серійних номерів у залишках'

    balance_id = fields.Many2one('stock.balance', 'Залишок', required=True)
    nomenclature_name = fields.Char('Товар', related='balance_id.nomenclature_id.name', readonly=True)
    location_info = fields.Char('Локація', compute='_compute_location_info', readonly=True)
    batch_info = fields.Char('Партія', compute='_compute_batch_info', readonly=True)
    
    # Кількості
    qty_on_hand = fields.Float('Фізична кількість', related='balance_id.qty_on_hand', readonly=True)
    qty_available = fields.Float('Доступна кількість', related='balance_id.qty_available', readonly=True)
    serial_numbers_count = fields.Integer('Введено S/N', related='balance_id.serial_numbers_count', readonly=True)
    missing_serial_count = fields.Float('Без S/N', related='balance_id.missing_serial_count', readonly=True)
    
    # Серійні номери
    serial_line_ids = fields.One2many('stock.balance.serial.detailed.wizard.line', 'wizard_id', 'Серійні номери')
    
    # Статистика
    has_missing_serials = fields.Boolean('Є товар без S/N', compute='_compute_statistics')
    completeness_percentage = fields.Float('% заповненості S/N', compute='_compute_statistics')

    @api.depends('balance_id')
    def _compute_location_info(self):
        for wizard in self:
            if wizard.balance_id.location_type == 'warehouse':
                wizard.location_info = f"Склад: {wizard.balance_id.warehouse_id.name}"
            else:
                wizard.location_info = f"Працівник: {wizard.balance_id.employee_id.name}"

    @api.depends('balance_id')
    def _compute_batch_info(self):
        for wizard in self:
            if wizard.balance_id.batch_id:
                batch = wizard.balance_id.batch_id
                wizard.batch_info = f"{batch.batch_number} (від {batch.date_created.strftime('%d.%m.%Y')})"
            else:
                wizard.batch_info = "Без партії"

    @api.depends('qty_on_hand', 'serial_numbers_count')
    def _compute_statistics(self):
        for wizard in self:
            if wizard.qty_on_hand > 0:
                wizard.completeness_percentage = (wizard.serial_numbers_count / wizard.qty_on_hand) * 100
                wizard.has_missing_serials = wizard.serial_numbers_count < wizard.qty_on_hand
            else:
                wizard.completeness_percentage = 0.0
                wizard.has_missing_serials = False

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        balance_id = self.env.context.get('default_balance_id')
        
        if balance_id:
            balance = self.env['stock.balance'].browse(balance_id)
            res['balance_id'] = balance_id
            
            # Отримуємо детальну інформацію про серійні номери
            detailed_serials = balance.get_serial_numbers_detailed()
            res['serial_line_ids'] = [(0, 0, {
                'sequence': serial['number'],
                'serial_number': serial['serial_number'],
                'status': serial['status'],
                'batch_number': serial['batch_number'],
                'location_name': serial['location'],
            }) for serial in detailed_serials]
        
        return res

    def action_export_serials(self):
        """Експортує серійні номери в Excel"""
        self.ensure_one()
        
        # Тут можна додати логіку експорту в Excel
        serials = [line.serial_number for line in self.serial_line_ids]
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Експорт'),
                'message': _('Експорт %d серійних номерів буде реалізовано') % len(serials),
                'type': 'info',
            }
        }

    def action_manage_serials(self):
        """Переходить до управління серійними номерами"""
        self.ensure_one()
        
        return {
            'name': f'Управління серійними номерами - {self.nomenclature_name}',
            'type': 'ir.actions.act_window',
            'res_model': 'stock.balance.serial.management.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_balance_id': self.balance_id.id},
        }


class StockBalanceSerialDetailedWizardLine(models.TransientModel):
    _name = 'stock.balance.serial.detailed.wizard.line'
    _description = 'Рядок детального перегляду серійного номера'
    _order = 'sequence, id'

    wizard_id = fields.Many2one('stock.balance.serial.detailed.wizard', 'Wizard', required=True, ondelete='cascade')
    sequence = fields.Integer('№', required=True)
    serial_number = fields.Char('Серійний номер', required=True)
    status = fields.Selection([
        ('available', 'Доступний'),
        ('reserved', 'Зарезервований'),
        ('consumed', 'Споживаний'),
    ], 'Статус', default='available')
    batch_number = fields.Char('Партія', readonly=True)
    location_name = fields.Char('Локація', readonly=True)


class StockBalanceSerialManagementWizard(models.TransientModel):
    _name = 'stock.balance.serial.management.wizard'
    _description = 'Управління серійними номерами в залишках'

    balance_id = fields.Many2one('stock.balance', 'Залишок', required=True)
    nomenclature_name = fields.Char('Товар', related='balance_id.nomenclature_id.name', readonly=True)
    location_info = fields.Char('Локація', compute='_compute_location_info', readonly=True)
    
    # Поточний стан
    qty_on_hand = fields.Float('Фізична кількість', related='balance_id.qty_on_hand', readonly=True)
    current_serials = fields.Text('Поточні серійні номери', related='balance_id.serial_numbers')
    
    # Редагування
    serial_line_ids = fields.One2many('stock.balance.serial.management.wizard.line', 'wizard_id', 'Серійні номери')
    
    # Додавання нових
    new_serials_text = fields.Text('Додати серійні номери', 
                                   help='Введіть серійні номери, розділені комою або новим рядком')
    
    # Статистика
    total_serials_count = fields.Integer('Загалом S/N', compute='_compute_counts')
    to_add_count = fields.Integer('Буде додано', compute='_compute_counts')
    to_remove_count = fields.Integer('Буде видалено', compute='_compute_counts')

    @api.depends('balance_id')
    def _compute_location_info(self):
        for wizard in self:
            if wizard.balance_id.location_type == 'warehouse':
                wizard.location_info = f"Склад: {wizard.balance_id.warehouse_id.name}"
            else:
                wizard.location_info = f"Працівник: {wizard.balance_id.employee_id.name}"

    @api.depends('serial_line_ids', 'new_serials_text')
    def _compute_counts(self):
        for wizard in self:
            existing_count = len([line for line in wizard.serial_line_ids if not line.to_delete])
            
            new_serials = []
            if wizard.new_serials_text:
                for line in wizard.new_serials_text.split('\n'):
                    for serial in line.split(','):
                        serial = serial.strip()
                        if serial:
                            new_serials.append(serial)
            
            wizard.total_serials_count = existing_count + len(new_serials)
            wizard.to_add_count = len(new_serials)
            wizard.to_remove_count = len([line for line in wizard.serial_line_ids if line.to_delete])

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        balance_id = self.env.context.get('default_balance_id')
        
        if balance_id:
            balance = self.env['stock.balance'].browse(balance_id)
            res['balance_id'] = balance_id
            
            serials = balance.get_serial_numbers_list()
            res['serial_line_ids'] = [(0, 0, {
                'serial_number': serial,
                'to_delete': False,
            }) for serial in serials]
        
        return res

    def action_save_changes(self):
        """Зберігає зміни в серійних номерах"""
        self.ensure_one()
        
        # Збираємо нові серійні номери
        new_serials = []
        
        # Додаємо існуючі (які не відмічені для видалення)
        for line in self.serial_line_ids:
            if not line.to_delete:
                new_serials.append(line.serial_number)
        
        # Додаємо нові з текстового поля
        if self.new_serials_text:
            for line in self.new_serials_text.split('\n'):
                for serial in line.split(','):
                    serial = serial.strip()
                    if serial and serial not in new_serials:
                        new_serials.append(serial)
        
        # Перевіряємо, чи не перевищує кількість серійних номерів фізичну кількість
        if len(new_serials) > self.qty_on_hand:
            raise ValidationError(
                _('Кількість серійних номерів (%d) не може перевищувати фізичну кількість (%d)') % 
                (len(new_serials), self.qty_on_hand)
            )
        
        # Оновлюємо залишок
        new_serial_text = '\n'.join(new_serials) if new_serials else ''
        self.balance_id.write({
            'serial_numbers': new_serial_text,
            'last_update': fields.Datetime.now(),
        })
        
        # Логуємо зміни
        self.balance_id.message_post(
            body=_('Оновлено серійні номери: %d S/N, доступна кількість: %d') % 
                 (len(new_serials), self.balance_id.qty_available),
            message_type='notification'
        )
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Збережено'),
                'message': _('Серійні номери успішно оновлено. Доступна кількість: %d') % len(new_serials),
                'type': 'success',
            }
        }

    def action_load_from_file(self):
        """Завантажує серійні номери з файлу"""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stock.balance.serial.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_management_wizard_id': self.id,
            }
        }


class StockBalanceSerialManagementWizardLine(models.TransientModel):
    _name = 'stock.balance.serial.management.wizard.line'
    _description = 'Рядок управління серійним номером'

    wizard_id = fields.Many2one('stock.balance.serial.management.wizard', 'Wizard', required=True, ondelete='cascade')
    serial_number = fields.Char('Серійний номер', required=True)
    to_delete = fields.Boolean('Видалити', default=False)


class StockBalanceSerialImportWizard(models.TransientModel):
    _name = 'stock.balance.serial.import.wizard'
    _description = 'Імпорт серійних номерів для залишків'
    
    management_wizard_id = fields.Many2one('stock.balance.serial.management.wizard', 'Management Wizard')
    file_data = fields.Binary('Excel файл', required=True)
    file_name = fields.Char('Назва файлу')
    replace_existing = fields.Boolean('Замінити існуючі', default=False,
                                     help='Якщо увімкнено, замінить всі існуючі серійні номери')
    
    def action_import(self):
        """Імпортує серійні номери з файлу"""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError('Будь ласка, оберіть файл!')
        
        try:
            import base64
            import xlrd
            from openpyxl import load_workbook
            import io
            
            file_content = base64.b64decode(self.file_data)
            file_name = self.file_name.lower() if self.file_name else ''

            serials = []
            if file_name.endswith('.xls'):
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                for row in range(sheet.nrows):
                    cell_value = sheet.cell_value(row, 0)
                    if cell_value:
                        if isinstance(cell_value, (int, float)) and cell_value.is_integer():
                            serials.append(str(int(cell_value)))
                        else:
                            serials.append(str(cell_value).strip())
            elif file_name.endswith('.xlsx'):
                workbook = load_workbook(filename=io.BytesIO(file_content))
                sheet = workbook.active
                for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                    cell_value = row[0]
                    if cell_value:
                        if isinstance(cell_value, (int, float)) and cell_value.is_integer():
                            serials.append(str(int(cell_value)))
                        else:
                            serials.append(str(cell_value).strip())
            else:
                raise UserError('Підтримуються лише файли форматів .xls або .xlsx!')

            # Очищаємо порожні серійні номери
            serials = [s for s in serials if s]
            
            # Оновлюємо wizard
            if self.replace_existing:
                # Замінюємо всі серійні номери
                self.management_wizard_id.new_serials_text = '\n'.join(serials)
                self.management_wizard_id.serial_line_ids = [(5, 0, 0)]  # Видаляємо всі існуючі
            else:
                # Додаємо до існуючих
                existing_text = self.management_wizard_id.new_serials_text or ''
                if existing_text:
                    new_text = existing_text + '\n' + '\n'.join(serials)
                else:
                    new_text = '\n'.join(serials)
                self.management_wizard_id.new_serials_text = new_text
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'stock.balance.serial.management.wizard',
                'view_mode': 'form',
                'res_id': self.management_wizard_id.id,
                'target': 'new',
                'context': {
                    'default_notification': f'Імпортовано {len(serials)} серійних номерів',
                }
            }
        
        except Exception as e:
            raise UserError(f'Помилка читання файлу: {str(e)}')